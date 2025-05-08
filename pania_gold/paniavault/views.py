from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404,reverse
from .forms import IndividualCustomerForm,CompanyCustomerForm,SupplierForm
from .forms import BuyRawInvoiceForm,BuyScrapInvoiceForm,ReciptMeltInvoiceForm,ReciptCraftInvoiceForm
from .models import BuyRawInvoice,BuyScrapInvoice,CompanyVault,ReciptMeltInvoice,ReciptCraftInvoice,Customer,Supplier
from vitrin.models import CraftPiece,OldPiece,CompanyVitrin
from meltvitrin.models import MeltPiece
from vitrin.forms import CraftPieceForm,OldPieceForm
from meltvitrin.forms import MeltPieceForm
import jdatetime
import time
import barcode
from barcode.writer import ImageWriter
from io import BytesIO
import openpyxl
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from django.db.models import Q
from django.db.models import Sum, F, ExpressionWrapper, IntegerField, DecimalField
from django.db.models.functions import Coalesce
from decimal import Decimal






# -----------------------------------
@login_required
def customer_register(request):
    customer_type = request.GET.get('customer_type')  # نوع مشتری از URL دریافت شود
    form = None
    if customer_type == 'individual':
        form = IndividualCustomerForm()
    elif customer_type == 'company':
        form = CompanyCustomerForm()
    if request.method == 'POST':
        if customer_type == 'individual':
            form = IndividualCustomerForm(request.POST)
        elif customer_type == 'company':
            form = CompanyCustomerForm(request.POST)

        if form.is_valid():
            customer = form.save(commit=False)  # ذخیره نشدن مستقیم
            customer.customer_type = customer_type  # تنظیم نوع مشتری

            try:
                if customer_type == 'company':
                    # بررسی تکراری بودن کد ملی برای کمپانی‌ها
                    if Customer.objects.filter(Q(melliserail=customer.melliserail) & Q(customer_type='company')).exists():
                        messages.error(request, "این شرکت قبلا ثبت شده است")
                        return render(request, 'paniavault/customer_register.html', {
                            'form': form,
                            'customer_type': customer_type,
                        })
                customer.full_clean()  # اجرای اعتبارسنجی
                customer.save()  # ذخیره در پایگاه داده
                messages.success(request, "مشتری با موفقیت ثبت شد")
                return redirect('paniavault:customer_register')
            except ValidationError as e:
                for field, errors in e.message_dict.items():
                    for error in errors:
                        messages.error(request, f"{error}")
    return render(request, 'paniavault/customer_register.html', {
        'form': form,
        'customer_type': customer_type,  # ارسال نوع مشتری به قالب
    })
# ----------------------------------------------
@login_required
def supplier_register(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            mellicode = form.cleaned_data.get('mellicode')
            if mellicode and Supplier.objects.filter(mellicode=mellicode).exists():
                messages.error(request, "تأمین‌کننده‌ای با این کد ملی قبلاً ثبت شده است.")
            else:
                form.save()
                messages.success(request, "تأمین‌کننده با موفقیت ثبت شد.")
                return redirect('paniavault:supplier_register')
    else:
        form = SupplierForm()

    return render(request, 'paniavault/supplier_register.html', {
        'form': form,
    })




# -======================= طلای خام RAW ============================

@login_required
def buyraw_invoice_register(request):
    vault, created = CompanyVault.objects.get_or_create(
        defaults={'company_balance': 0, 'company_assets': 0}
    )
    if request.method == 'POST':
        form = BuyRawInvoiceForm(request.POST)
        jalali_date = request.POST.get('invoice_date')
        if jalali_date:
            try:
                formatted_date = jdatetime.datetime.strptime(jalali_date, '%Y/%m/%d').strftime('%Y-%m-%d')
                form.data = form.data.copy()
                form.data['invoice_date'] = formatted_date
            except ValueError:
                messages.error(request, 'تاریخ وارد شده نامعتبر است')
                return render(request, 'paniavault/buyraw_invoice_register.html', {'form': form})
        if form.is_valid():
            supplier = form.cleaned_data.get('supplier')
            invoice_date = form.cleaned_data.get('invoice_date')
            net_weight = form.cleaned_data.get('net_weight')
            # بررسی وجود فاکتور مشابه
            existing_invoice = BuyRawInvoice.objects.filter(
                supplier=supplier,
                invoice_date=invoice_date,
                net_weight=net_weight
            ).first()
            if existing_invoice:
                messages.error(request, "این فاکتور قبلاً ثبت شده است")
                return render(request, 'paniavault/buyraw_invoice_register.html', {'form': form})
            raw_invoice = form.save(commit=False)  # ذخیره موقت
            raw_invoice.vault = vault
            raw_invoice.save()
            return redirect('paniavault:buyraw_invoice_list')  # تغییر مسیر به لیست فاکتورها
    else:
        form = BuyRawInvoiceForm()
    return render(request, 'paniavault/buyraw_invoice_register.html', {'form': form})


# --------------------------------------

@login_required
def buyraw_invoice_list(request):
    if request.GET:
        request.session['buyraw_invoice_list_filters'] = {
            'start_date': request.GET.get('start_date', ''),
            'end_date': request.GET.get('end_date', ''),
            'supply_name': request.GET.get('supply_name', ''),
            'supply_type': request.GET.getlist('supply_type', []),
        }

    filters = request.session.get('buyraw_invoice_list_filters', {})
    start_date = filters.get('start_date', '')
    end_date = filters.get('end_date', '')
    supply_name = filters.get('supply_name', '')
    supply_type = filters.get('supply_type', [])

    invoices = BuyRawInvoice.objects.all()
    start_date_jalali = ''
    end_date_jalali = ''

    if start_date and end_date:
        start_date = start_date.replace('/', '-')
        end_date = end_date.replace('/', '-')
        start_year, start_month, start_day = map(int, start_date.split('-'))
        end_year, end_month, end_day = map(int, end_date.split('-'))
        start_date_gregorian = jdatetime.date(start_year, start_month, start_day).togregorian()
        end_date_gregorian = jdatetime.date(end_year, end_month, end_day).togregorian()
        invoices = invoices.filter(invoice_date__range=(start_date_gregorian, end_date_gregorian))
        start_date_jalali = jdatetime.date.fromgregorian(date=start_date_gregorian).strftime('%Y/%m/%d')
        end_date_jalali = jdatetime.date.fromgregorian(date=end_date_gregorian).strftime('%Y/%m/%d')

    gold_balance = None
    if supply_name:
        invoices = invoices.filter(
            Q(supplier__first_name__icontains=supply_name) |
            Q(supplier__last_name__icontains=supply_name)
        )

        suppliers = Supplier.objects.filter(
            Q(first_name__icontains=supply_name) |
            Q(last_name__icontains=supply_name)
        )
        if suppliers.count() == 1:
            supplier = suppliers.first()
            # محاسبه مجموع خرید
            buy_total = BuyRawInvoice.objects.filter(supplier=supplier).aggregate(
                total_weight=Sum('net_weight')
            )['total_weight'] or Decimal('0.00')
            # محاسبه مجموع تحویل
            melt_total = ReciptMeltInvoice.objects.filter(supplier=supplier).aggregate(
                total_weight=Sum('net_weight')
            )['total_weight'] or Decimal('0.00')
            # بالانس طلا
            gold_balance = buy_total - melt_total

    if supply_type:
        invoices = invoices.filter(supply_type__in=supply_type)

    invoices = invoices.order_by('-invoice_date')
    if not (start_date or end_date or supply_name):
        invoices = invoices[:10]

    all_net_weight = invoices.aggregate(total_weight=Sum('net_weight'))['total_weight'] or 0
    all_invoice_price = invoices.aggregate(total_invoice_price=Sum('invoice_price'))['total_invoice_price'] or 0

    weighted_price_sum = invoices.aggregate(weighted_price=Sum(F('invoice_dailyprice') * F('net_weight')))
    total_weight = invoices.aggregate(total_weight=Sum('net_weight'))

    if total_weight['total_weight'] and weighted_price_sum['weighted_price']:
        weighted_avg_daily_price = int(weighted_price_sum['weighted_price'] / total_weight['total_weight'])
    else:
        weighted_avg_daily_price = 0

    invoices = invoices.annotate(
        total_paid=Coalesce(Sum('buyrawpayments__amount'), 0)
    ).annotate(
        remaining=ExpressionWrapper(
            F('total_paid') - F('invoice_price'),
            output_field=IntegerField()
        )
    )
    total_remaining = sum(invoice.remaining for invoice in invoices)

    context = {
        'invoices': invoices,
        'supply_name': supply_name,
        'filters': filters,
        'start_date': start_date_jalali,
        'end_date': end_date_jalali,
        'supply_type': supply_type,
        'all_net_weight': all_net_weight,
        'all_invoice_price': all_invoice_price,
        'weighted_avg_daily_price': weighted_avg_daily_price,
        'total_remaining': total_remaining,
        'gold_balance': gold_balance,
    }
    return render(request, 'paniavault/buyraw_invoice_list.html', context)


# ----------------------------------------------------
@login_required
def buyraw_invoice_detail(request, invoice_id):
    invoice = get_object_or_404(BuyRawInvoice, pk=invoice_id)
    payments = invoice.buyrawpayments.all()
    # جمع مبلغ پرداختی
    total_paid = sum(p.amount for p in payments)
    context = {
        'invoice': invoice,
        'payments': payments,
        'total_paid': total_paid,
    }
    return render(request, 'paniavault/buyraw_invoice_detail.html', context)


# ==========================    ثبت خرید طلای قیچیSCRAP   =========================


@login_required
def buyscrap_invoice_register(request):
    vault, created = CompanyVault.objects.get_or_create(
        defaults={'company_balance': 0, 'company_assets': 0}
    )
    if request.method == 'POST':
        form = BuyScrapInvoiceForm(request.POST)
        jalali_date = request.POST.get('invoice_date')
        if jalali_date:
            try:
                formatted_date = jdatetime.datetime.strptime(jalali_date, '%Y/%m/%d').strftime('%Y-%m-%d')
                form.data = form.data.copy()
                form.data['invoice_date'] = formatted_date
            except ValueError:
                messages.error(request, 'تاریخ وارد شده نامعتبر است')
                return render(request, 'paniavault/buyscrap_invoice_register.html', {'form': form})
        if form.is_valid():
            customer = form.cleaned_data.get('customer')
            invoice_date = form.cleaned_data.get('invoice_date')
            net_weight = form.cleaned_data.get('net_weight')
            # بررسی وجود فاکتور مشابه
            existing_invoice = BuyScrapInvoice.objects.filter(
                customer=customer,
                invoice_date=invoice_date,
                net_weight=net_weight
            ).first()
            if existing_invoice:
                messages.error(request, "این فاکتور قبلاً ثبت شده است")
                return render(request, 'paniavault/buyscrap_invoice_register.html', {'form': form})
            raw_invoice = form.save(commit=False)  # ذخیره موقت
            raw_invoice.vault = vault
            raw_invoice.save()
            return redirect('paniavault:buyscrap_invoice_list')  # تغییر مسیر به لیست فاکتورها
    else:
        form = BuyScrapInvoiceForm()
    return render(request, 'paniavault/buyscrap_invoice_register.html', {'form': form})


# --------------------------------------


@login_required
def buyscrap_invoice_list(request):
    if request.GET:
        request.session['buyscrap_invoice_list_filters'] = {
            'start_date': request.GET.get('start_date', ''),
            'end_date': request.GET.get('end_date', ''),
            'supply_name': request.GET.get('supply_name', ''),

        }

    filters = request.session.get('buyscrap_invoice_list_filters', {})
    start_date = filters.get('start_date', '')
    end_date = filters.get('end_date', '')
    supply_name = filters.get('supply_name', '')
    invoices = BuyScrapInvoice.objects.all()
    start_date_jalali = ''
    end_date_jalali = ''

    if start_date and end_date:
        start_date = start_date.replace('/', '-')
        end_date = end_date.replace('/', '-')
        start_year, start_month, start_day = map(int, start_date.split('-'))
        end_year, end_month, end_day = map(int, end_date.split('-'))
        start_date_gregorian = jdatetime.date(start_year, start_month, start_day).togregorian()
        end_date_gregorian = jdatetime.date(end_year, end_month, end_day).togregorian()
        invoices = invoices.filter(invoice_date__range=(start_date_gregorian, end_date_gregorian))
        start_date_jalali = jdatetime.date.fromgregorian(date=start_date_gregorian).strftime('%Y/%m/%d')
        end_date_jalali = jdatetime.date.fromgregorian(date=end_date_gregorian).strftime('%Y/%m/%d')

    if supply_name:
        invoices = invoices.filter(
            Q(customer__first_name__icontains=supply_name) |
            Q(customer__last_name__icontains=supply_name)
        )
    invoices = invoices.order_by('-invoice_date')
    if not (start_date or end_date or supply_name):
        invoices = invoices[:10]

    # محاسبه مجموع وزن خرید شده برای کل فاکتورها
    all_net_weight = invoices.aggregate(total_weight=Sum('net_weight'))['total_weight'] or 0
    all_invoice_price = invoices.aggregate(total_invoice_price=Sum('invoice_price'))['total_invoice_price'] or 0
    # محاسبه میانگین وزنی قیمت روز طلا
    weighted_price_sum = invoices.aggregate(weighted_price=Sum(F('invoice_dailyprice') * F('net_weight')))
    total_weight = invoices.aggregate(total_weight=Sum('net_weight'))

    # بررسی عدم تقسیم بر صفر و تبدیل به عدد صحیح
    if total_weight['total_weight'] and weighted_price_sum['weighted_price']:
        weighted_avg_daily_price = int(weighted_price_sum['weighted_price'] / total_weight['total_weight'])
    else:
        weighted_avg_daily_price = 0
    context = {
        'invoices': invoices,
        'supply_name': supply_name,
        'filters': filters,
        'start_date': start_date_jalali,
        'end_date': end_date_jalali,
        'all_net_weight': all_net_weight,
        'all_invoice_price': all_invoice_price,
        'weighted_avg_daily_price': weighted_avg_daily_price,
    }
    return render(request, 'paniavault/buyscrap_invoice_list.html', context)

# ---------------------------------------


@login_required
def company_vault_and_vitrin_list(request):
    vaults = CompanyVault.objects.all()
    vitrins = CompanyVitrin.objects.all()

    # فیلتر کردن آیتم‌هایی که فروخته نشده‌اند
    melt_pieces = MeltPiece.objects.filter(is_sold=False)
    craft_pieces = CraftPiece.objects.filter(is_sold=False)
    old_pieces = OldPiece.objects.filter(is_sold=False)

    # جمع کردن وزن خالص (net_weight) برای هر مدل
    total_melt_weight = melt_pieces.aggregate(total_weight=Sum('net_weight'))['total_weight'] or 0
    total_craft_weight = craft_pieces.aggregate(total_weight=Sum('net_weight'))['total_weight'] or 0
    total_old_weight = old_pieces.aggregate(total_weight=Sum('net_weight'))['total_weight'] or 0

    # جمع کل وزن‌ها
    total_net_weight = total_melt_weight + total_craft_weight + total_old_weight

    context = {
        'vaults': vaults,
        'vitrins': vitrins,
        'melt_pieces': melt_pieces,
        'craft_pieces': craft_pieces,
        'old_pieces': old_pieces,
        'total_melt_weight': total_melt_weight,
        'total_craft_weight': total_craft_weight,
        'total_old_weight': total_old_weight,
        'total_net_weight': total_net_weight,
    }
    return render(request, 'paniavault/company_vault_and_vitrin_list.html', context)

# ==============================  MELT ابشده   ===========================================
@login_required
def create_recipt_melt_invoice(request):
    if request.method == 'POST':
        form = ReciptMeltInvoiceForm(request.POST)
        jalali_date = request.POST.get('invoice_date')
        if jalali_date:
            try:
                formatted_date = jdatetime.datetime.strptime(jalali_date, '%Y/%m/%d').strftime('%Y-%m-%d')
                form.data = form.data.copy()
                form.data['invoice_date'] = formatted_date
            except ValueError:
                messages.error(request, 'تاریخ وارد شده نامعتبر است')
                return render(request, 'paniavault/create_recipt_melt_invoice.html', {'form': form})
        # بررسی صحت فرم
        if form.is_valid():
            invoice_date = form.cleaned_data.get('invoice_date')
            supplier = form.cleaned_data.get('supplier')
            net_weight = form.cleaned_data.get('net_weight')
            # بررسی وجود رکورد مشابه
            existing_invoice = ReciptMeltInvoice.objects.filter(
                invoice_date=invoice_date,
                supplier=supplier,
                net_weight=net_weight
            ).exists()
            if existing_invoice:
                messages.error(request, "این فاکتور قبلاً ثبت شده است")
                return render(request, 'paniavault/create_recipt_melt_invoice.html', {'form': form})

            invoice = form.save(commit=False)
            if not invoice.vault_id:
                default_vault = CompanyVault.objects.first()
                if not default_vault:
                    form.add_error(None, 'هیچ ولتی برای شرکت تعریف نشده است. لطفاً یک ولت اضافه کنید.')
                    return render(request, 'paniavault/create_recipt_melt_invoice.html', {
                        'form': form,
                    })
                invoice.vault = default_vault
            invoice.save()  # ذخیره فاکتور
            return redirect('paniavault:recipt_melt_invoice_list')
    else:
        form = ReciptMeltInvoiceForm()
    return render(request, 'paniavault/create_recipt_melt_invoice.html', {
        'form': form,
    })


# ------------ثبت خرید قطعه ابشده در فاکتور اصلی ------------------

@login_required
def register_melt_piece(request, invoice_id):
    invoice = get_object_or_404(ReciptMeltInvoice, id=invoice_id)
    pieces = MeltPiece.objects.filter(invoice=invoice)
    if request.method == 'POST':
        form = MeltPieceForm(request.POST)
        if form.is_valid():
            karat = form.cleaned_data['karat']
            weight = form.cleaned_data['weight']
            code = form.cleaned_data.get('code')
            if not code:  # اگر کد وجود ندارد
                if weight is not None:
                    int_part = int(weight)
                    dec_part = int((weight - int_part) * 100)
                    formatted_weight = f"{int_part:02d}{dec_part:02d}"
                else:
                    formatted_weight = "0000"
                timestamp = str(int(time.time() % 10))  # یک کاراکتر از تایم‌استمپ
                formatted_karat = f"{int(karat):03d}"[:3]  # سه کاراکتر از عیار
                # تولید کد با فرمت جدید
                generated_code = f"ME{timestamp}-{formatted_karat}-{formatted_weight}"
                form.instance.code = generated_code
            else:
                form.instance.code = code

            piece = form.save(commit=False)
            piece.invoice = invoice
            piece.save()
            messages.success(request, "قطعه با موفقیت ثبت شد.")
            return redirect(reverse('paniavault:register_melt_piece', args=[invoice_id]))
    else:
        form = MeltPieceForm()

    return render(request, 'paniavault/register_melt_piece.html', {
        'form': form,
        'invoice': invoice,
        'pieces': pieces,
    })


# ------------------لیست فاکتورهای خرید ابشده---------------------------

@login_required
def recipt_melt_invoice_list(request):
    if request.GET:
        request.session['recipt_melt_invoice_list_filters'] = {
            'start_date': request.GET.get('start_date', ''),
            'end_date': request.GET.get('end_date', ''),
            'supply_name': request.GET.get('supply_name', ''),

        }

    filters = request.session.get('recipt_melt_invoice_list_filters', {})
    start_date = filters.get('start_date', '')
    end_date = filters.get('end_date', '')
    supply_name = filters.get('supply_name', '')
    invoices = ReciptMeltInvoice.objects.all()
    start_date_jalali = ''
    end_date_jalali = ''

    if start_date and end_date:
        start_date = start_date.replace('/', '-')
        end_date = end_date.replace('/', '-')
        start_year, start_month, start_day = map(int, start_date.split('-'))
        end_year, end_month, end_day = map(int, end_date.split('-'))
        start_date_gregorian = jdatetime.date(start_year, start_month, start_day).togregorian()
        end_date_gregorian = jdatetime.date(end_year, end_month, end_day).togregorian()
        invoices = invoices.filter(invoice_date__range=(start_date_gregorian, end_date_gregorian))
        start_date_jalali = jdatetime.date.fromgregorian(date=start_date_gregorian).strftime('%Y/%m/%d')
        end_date_jalali = jdatetime.date.fromgregorian(date=end_date_gregorian).strftime('%Y/%m/%d')

    if supply_name:
        invoices = invoices.filter(
            Q(supplier__first_name__icontains=supply_name) |
            Q(supplier__last_name__icontains=supply_name)
        )
    invoices = invoices.order_by('-invoice_date')
    if not (start_date or end_date or supply_name):
        invoices = invoices[:10]

    # محاسبه مجموع وزن خرید شده برای کل فاکتورها
    all_net_weight = invoices.aggregate(total_weight=Sum('net_weight'))['total_weight'] or 0

    context = {
        'invoices': invoices,
        'supply_name': supply_name,
        'filters': filters,
        'start_date': start_date_jalali,
        'end_date': end_date_jalali,
        'all_net_weight': all_net_weight,
    }
    return render(request, 'paniavault/recipt_melt_invoice_list.html', context)


# ========================= CRAFT ==============================================

# ایجاد فاکتور خرید زینتی
@login_required
def create_recipt_craft_invoice(request):
    if request.method == 'POST':
        form = ReciptCraftInvoiceForm(request.POST)
        jalali_date = request.POST.get('invoice_date')
        if jalali_date:
            try:
                formatted_date = jdatetime.datetime.strptime(jalali_date, '%Y/%m/%d').strftime('%Y-%m-%d')
                form.data = form.data.copy()
                form.data['invoice_date'] = formatted_date
            except ValueError:
                messages.error(request, 'تاریخ وارد شده نامعتبر است.')
                return render(request, 'paniavault/create_recipt_craft_invoice.html', {'form': form})
        if form.is_valid():
            invoice_date = form.cleaned_data.get('invoice_date')
            supplier = form.cleaned_data.get('supplier')
            net_weight = form.cleaned_data.get('net_weight')
            # بررسی وجود رکورد مشابه
            existing_invoice = ReciptCraftInvoice.objects.filter(
                invoice_date=invoice_date,
                supplier=supplier,
                net_weight=net_weight
            ).exists()
            if existing_invoice:
                messages.error(request, "این فاکتور قبلاً ثبت شده است.")
                return render(request, 'paniavault/create_recipt_craft_invoice.html', {'form': form})
            invoice = form.save(commit=False)
            if not invoice.vault_id:
                default_vault = CompanyVault.objects.first()
                if not default_vault:
                    form.add_error(None, 'هیچ ولتی برای شرکت تعریف نشده است. لطفاً یک ولت اضافه کنید.')
                    return render(request, 'paniavault/create_recipt_craft_invoice.html', {
                        'form': form,
                    })
                invoice.vault = default_vault
            invoice.save()  # ذخیره فاکتور
            return redirect('paniavault:recipt_craft_invoice_list')
        else:
            messages.error(request, "لطفاً خطاهای موجود در فرم را بررسی کنید.")
    else:
        form = ReciptCraftInvoiceForm()
    return render(request, 'paniavault/create_recipt_craft_invoice.html', {
        'form': form,
    })

# ----------------------------لیست فاکتورهای زینتی---------------
@login_required
def recipt_craft_invoice_list(request):
    if request.GET:
        request.session['list_filters'] = {
            'start_date': request.GET.get('start_date', ''),
            'end_date': request.GET.get('end_date', ''),
            'supply_name': request.GET.get('supply_name', ''),
        }

    filters = request.session.get('list_filters', {})
    start_date = filters.get('start_date', '')
    end_date = filters.get('end_date', '')
    supply_name = filters.get('supply_name', '')

    invoices = ReciptCraftInvoice.objects.all()
    start_date_jalali = ''
    end_date_jalali = ''

    if start_date and end_date:
        start_date = start_date.replace('/', '-')
        end_date = end_date.replace('/', '-')
        start_year, start_month, start_day = map(int, start_date.split('-'))
        end_year, end_month, end_day = map(int, end_date.split('-'))
        start_date_gregorian = jdatetime.date(start_year, start_month, start_day).togregorian()
        end_date_gregorian = jdatetime.date(end_year, end_month, end_day).togregorian()
        invoices = invoices.filter(invoice_date__range=(start_date_gregorian, end_date_gregorian))
        start_date_jalali = jdatetime.date.fromgregorian(date=start_date_gregorian).strftime('%Y/%m/%d')
        end_date_jalali = jdatetime.date.fromgregorian(date=end_date_gregorian).strftime('%Y/%m/%d')

    if supply_name:
        invoices = invoices.filter(
            Q(supplier__first_name__icontains=supply_name) |
            Q(supplier__last_name__icontains=supply_name)
        )
    invoices = invoices.order_by('-invoice_date')
    if not (start_date or end_date or supply_name):
        invoices = invoices[:10]


    # محاسبه مجموع وزن خرید شده برای کل فاکتورها
    all_net_weight = invoices.aggregate(total_weight=Sum('net_weight'))['total_weight'] or 0
    all_invoice_price = invoices.aggregate(total_invoice_price=Sum('invoice_price'))['total_invoice_price'] or 0
    # محاسبه میانگین وزنی قیمت روز طلا
    weighted_price_sum = invoices.aggregate(weighted_price=Sum(F('invoice_dailyprice') * F('net_weight')))
    total_weight = invoices.aggregate(total_weight=Sum('net_weight'))

    # بررسی عدم تقسیم بر صفر و تبدیل به عدد صحیح
    if total_weight['total_weight'] and weighted_price_sum['weighted_price']:
        weighted_avg_daily_price = int(weighted_price_sum['weighted_price'] / total_weight['total_weight'])
    else:
        weighted_avg_daily_price = 0

    context = {
        'invoices': invoices,
        'supply_name': supply_name,
        'filters': filters,
        'start_date': start_date_jalali,
        'end_date': end_date_jalali,
        'all_net_weight': all_net_weight,
        'all_invoice_price': all_invoice_price,
        'weighted_avg_daily_price': weighted_avg_daily_price,
    }
    return render(request, 'paniavault/recipt_craft_invoice_list.html', context)

# ------------ثبت خرید قطعه زینتی در فاکتور اصلی ------------------

@login_required
def register_craft_piece(request, invoice_id):
    invoice = get_object_or_404(ReciptCraftInvoice, id=invoice_id)
    pieces = CraftPiece.objects.filter(invoice=invoice)
    if request.method == 'POST':
        form = CraftPieceForm(request.POST)
        if form.is_valid():
            gold_type = form.cleaned_data['gold_type']
            sale_ojrat = form.cleaned_data['sale_ojrat']
            net_weight = form.cleaned_data['net_weight']
            code = form.cleaned_data.get('code')
            if not code:  # اگر کد وجود ندارد
                if net_weight is not None:
                    int_part = int(net_weight)
                    dec_part = int((net_weight - int_part) * 100)
                    formatted_net_weight = f"{int_part:02d}{dec_part:02d}"
                else:
                    formatted_net_weight = "0000"
                timestamp = str(int(time.time() % 10))  # یک کاراکتر از تایم‌استمپ
                formatted_sale_ojrat = f"{int(sale_ojrat):02d}"[:2]  # سه کاراکتر از عیار
                first_letter = gold_type[:1].upper() if gold_type else ''
                generated_code = f"{first_letter}-{timestamp}-{formatted_sale_ojrat}-{formatted_net_weight}"
                form.instance.code = generated_code
            else:
                form.instance.code = code

            piece = form.save(commit=False)
            piece.invoice = invoice
            piece.save()
            messages.success(request, "قطعه با موفقیت ثبت شد.")
            return redirect(reverse('paniavault:register_craft_piece', args=[invoice_id]))
    else:
        form = CraftPieceForm()
    return render(request, 'paniavault/register_craft_piece.html', {
        'form': form,
        'invoice': invoice,
        'pieces': pieces,
    })

# --------------------دانلود اکسل لیست زینتی-----------------

@login_required
def download_craft_pieces_excel(request, invoice_id):
    invoice = get_object_or_404(ReciptCraftInvoice, id=invoice_id)
    pieces = CraftPiece.objects.filter(invoice=invoice)
    # ایجاد یک فایل اکسل جدید
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Craft Pieces"
    # افزودن عنوان ستون‌ها
    columns = ['ردیف', 'وزن', 'اجرت فروش', 'نام', 'کد', 'بارکد']
    for col_num, column_title in enumerate(columns, 1):
        sheet.cell(row=1, column=col_num, value=column_title)
    # تنظیم ارتفاع سلول‌ها به 20
    sheet.row_dimensions[1].height = 30  # تنظیم ارتفاع ردیف عنوان
    for row_num in range(2, len(pieces) + 2):
        sheet.row_dimensions[row_num].height = 30  # تنظیم ارتفاع ردیف‌های داده‌ها
    # افزودن داده‌های جدول همراه با بارکد
    for row_num, piece in enumerate(pieces, start=2):
        # اضافه کردن اطلاعات قطعه به جدول
        sheet.cell(row=row_num, column=1, value=row_num - 1)  # ردیف
        sheet.cell(row=row_num, column=2, value=piece.net_weight)
        sheet.cell(row=row_num, column=3, value=piece.sale_ojrat)
        sheet.cell(row=row_num, column=4, value=piece.name)
        sheet.cell(row=row_num, column=5, value=piece.code)
        # تولید بارکد
        barcode_buffer = BytesIO()
        barcode_class = barcode.get_barcode_class('code128')
        barcode_instance = barcode_class(piece.code, writer=ImageWriter())
        barcode_instance.write(barcode_buffer)
        # قرار دادن بارکد به عنوان تصویر در فایل اکسل
        barcode_buffer.seek(0)
        img = openpyxl.drawing.image.Image(barcode_buffer)
        img.width = 150  # تنظیم عرض تصویر بارکد
        img.height = 50  # تنظیم ارتفاع تصویر بارکد
        sheet.add_image(img, f'F{row_num}')  # اضافه کردن بارکد به ستون F
    # تنظیم پاسخ HTTP برای دانلود فایل
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="CraftPieces_{invoice_id}.xlsx"'
    workbook.save(response)
    return response
# -----------------------------------------

